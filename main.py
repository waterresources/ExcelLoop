import openpyxl as xl
import os

st_yr = 1980
sp_yr = 2021
num_of_years = sp_yr - st_yr

# create new workbook
location = 'C:/Users/Jason/Desktop/np_nitrate_analysis/data/final/'
print(f'Your working directory is: {location}')
processed_name = input('Name of Workbook (no extension): ').strip() + '.xlsx'
processed_workbook = xl.Workbook()
processed_workbook.save(location + processed_name)

# load workbook we created
processed_workbook = xl.load_workbook(location + processed_name)
processed_worksheet = processed_workbook.active

for year in range(num_of_years):

    raw_location = f'C:/Users/Jason/Desktop/np_nitrate_analysis/data/raw/{st_yr} Sampling Report.xlsx'
    raw_workbook = xl.load_workbook(raw_location)
    raw_worksheet = raw_workbook.active

    for row in raw_worksheet.iter_rows(min_row=2, max_row=raw_worksheet.max_row,
                                       min_col=1, max_col=3, values_only=True):
        processed_worksheet.append(row)

    st_yr += 1

# save final workbook
processed_workbook.save(location + processed_name)
print('Workbook Saved! Opening now')
os.startfile(location + processed_name)
